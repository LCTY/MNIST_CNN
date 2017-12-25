import os
import re
import math
import matplotlib.pyplot as plt
import numpy as np
import tensorflow as tf
from openpyxl import Workbook
from tensorflow.examples.tutorials.mnist import input_data
from fixedInt import *

# Select model.
model_name = "model_1"
print("Model name: %s" % model_name)

# Read MNIST database.
mnist = input_data.read_data_sets("MNIST_data/", one_hot=True)
print("Download Done!")

# Read model.
tf.reset_default_graph()
sess = tf.Session()
load_path = "data/" + model_name + "/weight/tf"
if not os.path.exists(load_path):
    raise IOError("Read file error.")
saver = tf.train.import_meta_graph(load_path + '/model.ckpt.meta')
saver.restore(sess, tf.train.latest_checkpoint(load_path))


# Truncation
def function_0():
    def truncation(x, trn):
        return tf.floor(x * (2 ** trn)) / (2 ** trn)

    def conv2d(x, W, conv2d_name=None):
        return tf.nn.conv2d(x, W, strides=[1, 1, 1, 1], padding='SAME', name=conv2d_name)

    def max_pool(x, max_pool_name=None):
        return tf.nn.max_pool(x, ksize=[1, 2, 2, 1], strides=[1, 2, 2, 1], padding='SAME', name=max_pool_name)

    for trn in range(8, -1, -1):
        # paras
        w_conv1 = tf.get_default_graph().get_tensor_by_name('w_conv1:0')
        b_conv1 = tf.get_default_graph().get_tensor_by_name('b_conv1:0')

        # conv layer-1
        x = tf.placeholder(tf.float32, [None, 784], name="x")
        _x = truncation(x, 3)
        x_image = tf.reshape(_x, [-1, 28, 28, 1])

        h_conv1 = tf.nn.relu(conv2d(x_image, w_conv1, 'conv2d1') + b_conv1, name="h_conv1")
        _h_conv1 = truncation(h_conv1, 4)
        h_pool1 = max_pool(_h_conv1, 'max_pool1')

        # full connection
        w_fc1 = tf.get_default_graph().get_tensor_by_name('w_fc1:0')
        b_fc1 = tf.get_default_graph().get_tensor_by_name('b_fc1:0')

        h_pool1_flat = tf.reshape(h_pool1, [-1, 14 * 14 * 10])
        h_fc1 = tf.nn.relu(tf.matmul(h_pool1_flat, w_fc1) + b_fc1, name="h_fc1")
        _h_fc1 = truncation(h_fc1, trn)

        # dropout
        keep_prob = tf.placeholder(tf.float32, name="keep_prob")
        h_fc1_drop = tf.nn.dropout(_h_fc1, keep_prob, name="h_fc1_drop")

        # output layer: softmax
        w_fc2 = tf.get_default_graph().get_tensor_by_name('w_fc2:0')
        b_fc2 = tf.get_default_graph().get_tensor_by_name('b_fc2:0')

        y_conv = tf.nn.softmax(tf.matmul(h_fc1_drop, w_fc2) + b_fc2, name="y_conv")
        y_ = tf.placeholder(tf.float32, [None, 10], name="y_")

        correct_prediction = tf.equal(tf.argmax(y_conv, 1), tf.argmax(y_, 1))
        accuracy = tf.reduce_mean(tf.cast(correct_prediction, tf.float32))

        train_accuracy = sess.run(accuracy, feed_dict={x: mnist.test.images, y_: mnist.test.labels, keep_prob: 1.0})
        print('Truncation bits: {:d}, Mean accuracy: {:.8f}'.format(trn, train_accuracy))


# 算出test資料集的精準度
def function_1():
    x = tf.get_default_graph().get_tensor_by_name('x:0')
    y_ = tf.get_default_graph().get_tensor_by_name('y_:0')
    keep_prob = tf.get_default_graph().get_tensor_by_name('keep_prob:0')
    y_conv = tf.get_default_graph().get_tensor_by_name('y_conv:0')
    x_test, y_test = mnist.test.images, mnist.test.labels

    correct_prediction = tf.equal(tf.argmax(y_conv, 1), tf.argmax(y_, 1))
    accuracy = tf.reduce_mean(tf.cast(correct_prediction, tf.float32))
    train_accuracy = sess.run(accuracy, feed_dict={x: x_test, y_: y_test, keep_prob: 1.0})
    print('Mean accuracy: {:.8f}'.format(train_accuracy))


# 輸出activation value分布直方圖(only for tf_model)
def function_2():
    # 用於找出各層名字
    # op = sess.graph.get_operations()
    # tf.get_default_graph().get_tensor_by_name('Reshape_1:0')
    # for idx, val in enumerate(op):
    #     print(val)

    x = tf.get_default_graph().get_tensor_by_name('x:0')
    y_ = tf.get_default_graph().get_tensor_by_name('y_:0')
    keep_prob = tf.get_default_graph().get_tensor_by_name('keep_prob:0')

    if model_name == "model_1" or model_name == "model_3":
        h_conv1 = tf.get_default_graph().get_tensor_by_name('h_conv1:0')
        h_pool1 = tf.get_default_graph().get_tensor_by_name('max_pool1:0')
        h_conv2 = tf.get_default_graph().get_tensor_by_name('h_conv2:0')
        h_pool2 = tf.get_default_graph().get_tensor_by_name('max_pool2:0')
        h_fc1 = tf.get_default_graph().get_tensor_by_name('h_fc1:0')
        y_conv = tf.get_default_graph().get_tensor_by_name('y_conv:0')
        input_tuple = h_conv1, h_pool1, h_conv2, h_pool2, h_fc1, y_conv
    else:
        h_conv1 = tf.get_default_graph().get_tensor_by_name('h_conv1:0')
        h_pool1 = tf.get_default_graph().get_tensor_by_name('max_pool1:0')
        h_fc1 = tf.get_default_graph().get_tensor_by_name('h_fc1:0')
        y_conv = tf.get_default_graph().get_tensor_by_name('y_conv:0')
        input_tuple = h_conv1, h_pool1, h_fc1, y_conv

    for index, value in enumerate(input_tuple):
        weight = sess.run(value, feed_dict={x: mnist.test.images[0:1], y_: mnist.test.labels[0:1], keep_prob: 1.0})
        weight = weight.flatten()

        for idx, val in enumerate(weight):
            if val == 0:
                weight[idx] = 0
            else:
                weight[idx] = math.log(abs(val), 2)

        plt.figure(figsize=(20, 10), dpi=300)
        plt.title(value.name)
        plt.xticks(range(-12, 13, 1))
        plt.hist(weight, bins=range(-12, 13, 1))
        save_path = "data/" + model_name + "/hist/activation/"
        if not os.path.exists(save_path):
            os.makedirs(save_path)
        plt.savefig(save_path + re.sub(':0', '', value.name) + ".png")
        plt.close()


# 輸出weight分布直方圖
def function_3():
    all_vars = tf.trainable_variables()
    for idx, val in enumerate(all_vars):
        # if idx == 1:
        #     break
        sv = tf.reshape(val, [-1])
        weight = sess.run(sv)
        for index, value in enumerate(weight):
            weight[index] = math.log(abs(value), 2)

        plt.figure(figsize=(20, 10), dpi=300)
        plt.title(val.name)
        plt.xticks(range(-12, 13, 1))
        plt.hist(weight, bins=range(-12, 13, 1))
        save_path = "data/" + model_name + "/hist/weight/"
        if not os.path.exists(save_path):
            os.makedirs(save_path)
        plt.savefig(save_path + re.sub(':0', '', val.name) + ".png")
        plt.close()


# Convert weights of tf_model to hipsternet_model.
def function_4():
    def choose_model(index, x):
        if model_name == "model_1":
            if index == 0:
                return np.transpose(x, (3, 2, 0, 1))
            if index == 1:
                return np.reshape(x, (32, 1))
            if index == 2:
                return np.transpose(x, (3, 2, 0, 1))
            if index == 3:
                return np.reshape(x, (64, 1))
            if index == 4:
                return np.reshape(x, (3136, 1024))
            if index == 5:
                return np.reshape(x, (1, 1024))
            if index == 6:
                return np.reshape(x, (1024, 10))
            if index == 7:
                return np.reshape(x, (1, 10))
        elif model_name == "model_2":
            if index == 0:
                return np.transpose(x, (3, 2, 0, 1))
            if index == 1:
                return np.reshape(x, (10, 1))
            if index == 2:
                return np.reshape(x, (1960, 128))
            if index == 3:
                return np.reshape(x, (1, 128))
            if index == 4:
                return np.reshape(x, (128, 10))
            if index == 5:
                return np.reshape(x, (1, 10))
        elif model_name == "model_3":
            if index == 0:
                return np.transpose(x, (3, 2, 0, 1))
            if index == 1:
                return np.reshape(x, (10, 1))
            if index == 2:
                return np.transpose(x, (3, 2, 0, 1))
            if index == 3:
                return np.reshape(x, (20, 1))
            if index == 4:
                return np.reshape(x, (980, 128))
            if index == 5:
                return np.reshape(x, (1, 128))
            if index == 6:
                return np.reshape(x, (128, 10))
            if index == 7:
                return np.reshape(x, (1, 10))

    all_vars = tf.trainable_variables()
    for idx, val in enumerate(all_vars):
        # if idx == 1:
        #     break
        weight = sess.run(val)

        # print(weight.shape)
        weight = choose_model(idx, weight)
        # print(weight.shape)

        save_path = "data/" + model_name + "/weight/hipsternet/"
        if not os.path.exists(save_path):
            os.makedirs(save_path)
        save_path = save_path + re.sub(':0', '', val.name)
        np.save(save_path, weight)


# 將weight輸出成excel資料
def function_5():
    wb = Workbook()
    ws = wb.active

    all_vars = tf.trainable_variables()
    for idx, val in enumerate(all_vars):
        # if idx == 1:
        #     break

        ws.cell(row=1, column=idx + 1, value=val.name)

        sv = tf.reshape(val, [-1])
        weight = sess.run(sv)

        for index, value in enumerate(weight):
            ws.cell(row=index + 2, column=idx + 1, value=value)

    wb.save(load_path + "/weight.xlsx")


# 輸出activation value分布直方圖(only for tf_model)
def function_6():
    # 用於找出各層名字
    # op = sess.graph.get_operations()
    # tf.get_default_graph().get_tensor_by_name('Reshape_1:0')
    # for idx, val in enumerate(op):
    #     print(val)

    x = tf.get_default_graph().get_tensor_by_name('x:0')
    y_ = tf.get_default_graph().get_tensor_by_name('y_:0')
    keep_prob = tf.get_default_graph().get_tensor_by_name('keep_prob:0')

    if model_name == "model_1" or model_name == "model_3":
        h_conv1 = tf.get_default_graph().get_tensor_by_name('h_conv1:0')
        h_pool1 = tf.get_default_graph().get_tensor_by_name('max_pool1:0')
        h_conv2 = tf.get_default_graph().get_tensor_by_name('h_conv2:0')
        h_pool2 = tf.get_default_graph().get_tensor_by_name('max_pool2:0')
        h_fc1 = tf.get_default_graph().get_tensor_by_name('h_fc1:0')
        y_conv = tf.get_default_graph().get_tensor_by_name('y_conv:0')
        input_tuple = h_conv1, h_pool1, h_conv2, h_pool2, h_fc1, y_conv
    else:
        h_conv1 = tf.get_default_graph().get_tensor_by_name('h_conv1:0')
        h_pool1 = tf.get_default_graph().get_tensor_by_name('max_pool1:0')
        h_fc1 = tf.get_default_graph().get_tensor_by_name('h_fc1:0')
        y_conv = tf.get_default_graph().get_tensor_by_name('y_conv:0')
        input_tuple = h_conv1, h_pool1, h_fc1, y_conv

    for index, value in enumerate(input_tuple):
        weight = sess.run(value, feed_dict={x: mnist.test.images[0:1], y_: mnist.test.labels[0:1], keep_prob: 1.0})
        if index <= 3:
            weight = np.transpose(weight, (0, 3, 1, 2))
        weight = weight.flatten()

        # Save the variables to disk.
        save_path = "data/" + model_name + "/act/"
        if not os.path.exists(save_path):
            os.makedirs(save_path)
        np.savetxt(save_path + value.name + ".txt", weight, fmt='%.8e')


# function_0()
function_1()
# function_2()
# function_3()
# function_4()
# function_5()
# function_6()
